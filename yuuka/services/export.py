"""
Export service for ledger data.

Provides functionality to export ledger entries to XLSX and CSV formats.
"""

import csv
import io
from datetime import date, datetime
from enum import Enum
from typing import Optional, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from yuuka.db import LedgerEntry, LedgerRepository


class ExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    XLSX = "xlsx"


class ExportService:
    """Service for exporting ledger data to various formats."""

    def __init__(self, repository: LedgerRepository):
        """
        Initialize the export service.

        Args:
            repository: Repository for ledger entries
        """
        self.repository = repository

    def export_to_csv(
        self,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> io.BytesIO:
        """
        Export ledger entries to CSV format.

        Args:
            user_id: Discord user ID
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            BytesIO buffer containing the CSV data
        """
        entries = self._get_entries(user_id, start_date, end_date)

        buffer = io.BytesIO()
        text_buffer = io.StringIO()

        writer = csv.writer(text_buffer)

        # Write header
        writer.writerow(
            [
                "ID",
                "Date",
                "Time",
                "Action",
                "Amount",
                "Source",
                "Destination",
                "Description",
                "Raw Text",
                "Confidence",
            ]
        )

        # Write data rows
        for entry in entries:
            writer.writerow(
                [
                    entry.id,
                    entry.created_at.strftime("%Y-%m-%d"),
                    entry.created_at.strftime("%H:%M:%S"),
                    entry.action.value,
                    entry.amount,
                    entry.source or "",
                    entry.destination or "",
                    entry.description or "",
                    entry.raw_text,
                    f"{entry.confidence:.2f}",
                ]
            )

        # Convert to bytes
        buffer.write(text_buffer.getvalue().encode("utf-8-sig"))  # BOM for Excel
        buffer.seek(0)

        return buffer

    def export_to_xlsx(
        self,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> io.BytesIO:
        """
        Export ledger entries to XLSX format with formatting.

        Args:
            user_id: Discord user ID
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            BytesIO buffer containing the XLSX data
        """
        entries = self._get_entries(user_id, start_date, end_date)

        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = "Ledger"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        incoming_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        outgoing_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        transfer_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )

        # Headers
        headers = [
            "ID",
            "Date",
            "Time",
            "Action",
            "Amount",
            "Source",
            "Destination",
            "Description",
            "Raw Text",
            "Confidence",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, entry in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=entry.id)
            ws.cell(row=row_idx, column=2, value=entry.created_at.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=3, value=entry.created_at.strftime("%H:%M:%S"))
            ws.cell(row=row_idx, column=4, value=entry.action.value)
            ws.cell(row=row_idx, column=5, value=entry.amount)
            ws.cell(row=row_idx, column=6, value=entry.source or "")
            ws.cell(row=row_idx, column=7, value=entry.destination or "")
            ws.cell(row=row_idx, column=8, value=entry.description or "")
            ws.cell(row=row_idx, column=9, value=entry.raw_text)
            ws.cell(row=row_idx, column=10, value=entry.confidence)

            # Color code by action type
            action = entry.action.value
            if action == "incoming":
                fill = incoming_fill
            elif action == "outgoing":
                fill = outgoing_fill
            else:
                fill = transfer_fill

            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = fill

        # Format amount column as number
        for row in range(2, len(entries) + 2):
            ws.cell(row=row, column=5).number_format = "#,##0.00"
            ws.cell(row=row, column=10).number_format = "0.00"

        # Auto-adjust column widths
        column_widths = [8, 12, 10, 10, 15, 15, 15, 20, 40, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary sheet
        self._add_summary_sheet(wb, entries, user_id)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _add_summary_sheet(
        self,
        wb: Workbook,
        entries: list[LedgerEntry],
        user_id: str,
    ):
        """Add a summary sheet to the workbook."""
        ws = wb.create_sheet(title="Summary")

        # Styles
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)

        # Title
        ws.cell(row=1, column=1, value="Ledger Summary").font = title_font
        ws.cell(
            row=2,
            column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        # Calculate totals
        total_incoming = sum(e.amount for e in entries if e.action.value == "incoming")
        total_outgoing = sum(e.amount for e in entries if e.action.value == "outgoing")
        total_transfer = sum(e.amount for e in entries if e.action.value == "transfer")
        net = total_incoming - total_outgoing

        # Summary table
        summary_start = 4
        ws.cell(row=summary_start, column=1, value="Category").font = header_font
        ws.cell(row=summary_start, column=2, value="Count").font = header_font
        ws.cell(row=summary_start, column=3, value="Total").font = header_font

        incoming_count = sum(1 for e in entries if e.action.value == "incoming")
        outgoing_count = sum(1 for e in entries if e.action.value == "outgoing")
        transfer_count = sum(1 for e in entries if e.action.value == "transfer")

        ws.cell(row=summary_start + 1, column=1, value="Incoming")
        ws.cell(row=summary_start + 1, column=2, value=incoming_count)
        ws.cell(row=summary_start + 1, column=3, value=total_incoming)

        ws.cell(row=summary_start + 2, column=1, value="Outgoing")
        ws.cell(row=summary_start + 2, column=2, value=outgoing_count)
        ws.cell(row=summary_start + 2, column=3, value=total_outgoing)

        ws.cell(row=summary_start + 3, column=1, value="Transfer")
        ws.cell(row=summary_start + 3, column=2, value=transfer_count)
        ws.cell(row=summary_start + 3, column=3, value=total_transfer)

        ws.cell(row=summary_start + 5, column=1, value="Net Balance").font = header_font
        ws.cell(row=summary_start + 5, column=3, value=net)

        # Format numbers
        for row in range(summary_start + 1, summary_start + 6):
            ws.cell(row=row, column=3).number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 18

    def _get_entries(
        self,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> list[LedgerEntry]:
        """
        Get ledger entries with optional date filtering.

        Args:
            user_id: Discord user ID
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            List of LedgerEntry objects
        """
        if start_date and end_date:
            return self.repository.get_entries_for_date_range(
                user_id, start_date, end_date
            )
        else:
            # Get all entries (with a high limit)
            return self.repository.get_user_entries(user_id, limit=10000, offset=0)

    def get_filename(
        self,
        user_id: str,
        format: ExportFormat,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> str:
        """
        Generate a filename for the export.

        Args:
            user_id: Discord user ID
            format: Export format
            start_date: Optional start date
            end_date: Optional end date

        Returns:
            Suggested filename
        """
        date_str = datetime.now().strftime("%Y%m%d")

        if start_date and end_date:
            date_range = (
                f"_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"
            )
        else:
            date_range = ""

        return f"yuuka_ledger_{date_str}{date_range}.{format.value}"
